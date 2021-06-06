import xlrd
from openpyxl import Workbook, styles, load_workbook
from Tool.tool import *


class rExcel:
    # 打开excel文件
    def openExcle(self, path):
        wb = xlrd.open_workbook(path)
        return wb

    # 获取所有的sheet表
    def getSheetNames(self, workObj):
        sheetNames = workObj.sheet_names()
        return sheetNames

    # 切换活动表
    def checkOutSheet(self, workObj, sheet_name):
        sheet = workObj.sheet_by_name(sheet_name)
        return sheet

    # 获取所有行的数据
    def getRowData(self, sheetObj):
        list_data = []
        for i in range(sheetObj.nrows):
            list_data.append(sheetObj.row_values(i))
        return list_data


class wExcle:

    # 新建文件
    def creatExcel(self):
        newwb = Workbook()
        newwb.save('../result/植入单-中南大学湘雅三医院.xlsx')

    # 打开文件
    def openExcel(self, path):
        wb = load_workbook(path)
        return wb

    # 获取所有的sheet名
    def getSheetName(self, workObj):
        sheet_names = workObj.sheetnames
        return sheet_names

    # 切换sheet表
    def checkOutSheet(self, workObj, sheetName):
        sheet = workObj[sheetName]
        return sheet

    # 新建sheet表
    def creatSheet(self, wbObj):
        wbObj.create_sheet(realTime(), 0)
        sheet = wbObj.active
        return sheet

    # 创建模板表头
    def newRemplate(self, sheetobj):
        list_1 = ['序号', '手术日期', '患者姓名', '患者年龄', '住院号', '手术医师', '区', '床位号', '产品名称（实际使用）', '产品描述', '规格型号（实际使用）', '批号',
                  '储存条件', '生产日期', '失效日期', '注册证编号', '生产厂商', '生产许可证号', '数量（实际使用）', '单价（底价）', '底价成本', '产品编码（医院计费）',
                  '产品名称（医院计费）', '规格型号', '产品批号', '数量', '单价', '小计', '订单金额', '华润7%', '销售额', '跟台费', '消毒费', '餐费', '业务提成',
                  '台数', '临床30%', '分销总价', '分销利润', '分销单价']
        fille = styles.PatternFill("solid", fgColor="2894FF")
        for i in range(len(list_1)):
            sheetobj.cell(row=1, column=i + 1, value=list_1[i]).fill = fille


if __name__ == '__main__':
    open = rExcel()
    wb = open.openExcle("../DATA/2021-06/记账.xls")
    sheet = open.checkOutSheet(wb, 'Sheet1')
    col = sheet.ncols
    row = sheet.nrows
    print(col, row)
    for i in range(row):
        print(sheet.row_values(i))

# newwb.save("../result/植入单-中南大学湘雅三医院{}.xlsx".format(realTime()))
